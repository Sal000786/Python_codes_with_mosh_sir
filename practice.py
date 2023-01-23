import openpyxl as xl
from openpyxl.chart import BarChart,Reference
wb=xl.load_workbook('./Work_with_Excel_Spreadsheet/practice_data.xlsx')
# print(wb.sheetnames)
sheet=wb['Sheet1']
cell=sheet['a1']
print(cell.value)

print("max row are ",sheet.max_row)



for row in range(2,sheet.max_row+1):
    cell=sheet.cell(row,3)
    print(cell.value)

    new_price_cal=cell.value*0.9
    new_price_cell=sheet.cell(row,4)
    new_price_cell.value=new_price_cal
    new_price_cell_name=sheet.cell(row=1,column=4)
    new_price_cell_name="updated_price"


values=Reference(sheet,min_row=2,max_row=sheet.max_row,min_col=4,max_col=4)
chart=BarChart()
chart.add_data(values)
sheet.add_chart(chart,'e2')

wb.save('./Work_with_Excel_Spreadsheet/practice_data2.xlsx')
print('finished')