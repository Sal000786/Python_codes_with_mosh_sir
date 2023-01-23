import openpyxl as xl
wb=xl.load_workbook('./Work_with_Excel_Spreadsheet/transection.xlsx')
sheet=wb['Sheet1']
cell=sheet['a1']
# cell=sheet.cell(1,1)
print("cell value is ",cell.value)


print("maximum rows in the sheet are ",sheet.max_row)


#iterate through the rows

for rows in range(2,sheet.max_row +1):
    cell=sheet.cell(rows,3)
    # print(cell.value)
    corrected_value=cell.value *0.9
    # print(corrected_value)
    corrected_value_cell=sheet.cell(rows,4)
    corrected_value_cell.value=corrected_value
wb.save('./Work_with_Excel_Spreadsheet/transection2.xlsx')
# wb.remove('transection2.xlsx')
print('finished')
